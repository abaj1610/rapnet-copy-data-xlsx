import openpyxl
import streamlit as st

def find_last_non_empty_row(worksheet):
    max_row = worksheet.max_row
    for row_num in range(max_row, 0, -1):
        if any(worksheet.cell(row=row_num, column=col_num).value for col_num in range(1, worksheet.max_column + 1)):
            return row_num + 1  # Return the next row after the last non-empty row
    return 1  # If the sheet is empty, start from the first row

def append_elements_to_excel(input_string, columns_data, file_path, sheet_name):
     # Open the existing Excel file
    workbook = openpyxl.load_workbook(file_path)

    # Select the active worksheet
    worksheet = workbook[sheet_name]

    # Find the next empty row in the worksheet
    next_empty_row = find_last_non_empty_row(worksheet)

    # Split the input string into a list
    string_list = input_string.split('\n')

    # Write the specified elements of the list to specified columns in the Excel sheet
    for column, element in columns_data.items():
        if element < len(string_list):
            value = string_list[element]
            if isinstance(value, str) and value.startswith('$'):
                    value = float(value[1:])
            if isinstance(value, str) and value.endswith('%'):
                    value = float(value[:-1])
            worksheet.cell(row=next_empty_row, column=column, value=value)

    # print(string_list)
    # Save the workbook
    workbook.save(file_path)
    
    # print("Data appended to Excel file successfully!")


def main():

    # Streamlit Title
    st.title('Website for Copying Rapnet Data')
    
    # A key for the text area to manage its content via session state
    TEXT_AREA_KEY = 'user_input'

    # Check if the form has been submitted (using a flag in the session state)
    if 'form_submitted' not in st.session_state:
        st.session_state['form_submitted'] = False

    # If the form was just submitted, clear the text area and reset the submission flag
    if st.session_state['form_submitted']:
        st.session_state[TEXT_AREA_KEY] = ""  # Clear the text area
        st.session_state['form_submitted'] = False  # Reset the flag
        st.write("Copied the Data Successfully!")
    
    # Start a Streamlit Form
    with st.form(key='my_form'):
        input_string = st.text_area(label='Enter the Copied Data:',value=st.session_state.get(TEXT_AREA_KEY, ''), key=TEXT_AREA_KEY)
        submit_button = st.form_submit_button(label='Submit')
    
    # Map each column index to the corresponding element index in the input string
    columns_data = {3: -2, 9: 22, 10: 21, 11: 4, 12: 5, 13: 6, 15: 7, 16: 8, 17: 9, 18: 10, 19: 11, 20: 20, 21: 18, 22: 19, 25: 13, 26:15, 24: 14}

    file_path = "Buying Data.xlsx"
    sheet_name = "B & S"
    
    # This block runs only when the submit button is clicked
    if submit_button:
        # Calling the Function of Converting the Data into a CSV
        append_elements_to_excel(input_string, columns_data, file_path, sheet_name)
        st.session_state['form_submitted'] = True  # Set the flag indicating a form submission
        st.rerun()  # Rerun the script immediately to reflect change

main()

